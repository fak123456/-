"""CLI entrypoint: ``python -m crawler.run --input urls.xlsx --out crawler/output``.

Also supports ``--make-template`` to (re)generate the 1-column input
template at the path given via ``--input``.
"""

from __future__ import annotations

import argparse
import logging
import sys
from pathlib import Path


def _force_utf8_console() -> None:
    """Make stdout/stderr emit UTF-8 even when the system codepage is GBK.

    Required for the frozen exe on Chinese Windows: product titles may
    contain characters (e.g. U+200B zero-width space) that the default
    cp936 stdout cannot encode, which would crash the run.
    """
    for stream in (sys.stdout, sys.stderr):
        try:
            stream.reconfigure(encoding="utf-8", errors="replace")
        except (AttributeError, ValueError):
            pass


_force_utf8_console()

from openpyxl import Workbook

from crawler.runner import run_batch


def _build_parser() -> argparse.ArgumentParser:
    p = argparse.ArgumentParser(
        prog="python -m crawler.run",
        description=(
            "Standalone Amazon product crawler. Reads URLs from column A of "
            "an xlsx file and writes one {ASIN}.zip per product plus a "
            "generator-friendly 商品列表.xlsx into --out."
        ),
    )
    p.add_argument("--input", required=True, type=Path,
                   help="Input .xlsx whose column A contains product URLs.")
    p.add_argument("--out", default=Path("crawler/output"), type=Path,
                   help="Output directory (default: crawler/output).")
    p.add_argument("--max-images", type=int, default=8,
                   help="Maximum images per product (default: 8).")
    p.add_argument("--limit", type=int, default=None,
                   help="Process only the first N URLs (debug).")
    p.add_argument("--skip-existing", action="store_true",
                   help="Skip ASINs whose zip already exists in --out (resume after crash).")
    p.add_argument("--delay-min", type=float, default=2.0,
                   help="Minimum random delay (s) before each page fetch (default: 2.0).")
    p.add_argument("--delay-max", type=float, default=5.0,
                   help="Maximum random delay (s) before each page fetch (default: 5.0).")
    p.add_argument("--make-template", action="store_true",
                   help="Write a blank 1-column input template at --input and exit.")
    p.add_argument("-v", "--verbose", action="store_true",
                   help="Enable DEBUG-level logging.")
    return p


def _make_template(path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "URLs"
    ws.cell(1, 1, "商品链接")
    ws.cell(2, 1, "https://www.amazon.de/dp/B0GK22CBCR")
    ws.column_dimensions["A"].width = 70
    wb.save(str(path))
    print(f"Wrote template -> {path}")


def main(argv: list[str] | None = None) -> int:
    args = _build_parser().parse_args(argv)
    logging.basicConfig(
        level=logging.DEBUG if args.verbose else logging.INFO,
        format="%(asctime)s [%(levelname)s] %(name)s: %(message)s",
    )

    if args.make_template:
        _make_template(args.input)
        return 0

    if not args.input.is_file():
        print(f"ERROR: input file not found: {args.input}", file=sys.stderr)
        return 2

    if args.delay_min < 0 or args.delay_max < args.delay_min:
        print("ERROR: --delay-min/--delay-max must satisfy 0 <= min <= max",
              file=sys.stderr)
        return 2

    counts = run_batch(
        input_xlsx=args.input,
        out_dir=args.out,
        max_images=args.max_images,
        limit=args.limit,
        skip_existing=args.skip_existing,
        delay_min=args.delay_min,
        delay_max=args.delay_max,
    )
    # Exit code: 0 if at least one product succeeded, 1 otherwise.
    return 0 if counts.get("ok", 0) > 0 else 1


if __name__ == "__main__":
    raise SystemExit(main())

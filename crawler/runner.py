"""Top-level batch driver: read input xlsx column A, loop ``process_one``,
write a generator-friendly ``商品列表.xlsx`` plus a ``crawl_log.txt`` audit
trail.

The output ``商品列表.xlsx`` schema (`ZIP/文件夹路径` | `商品标题`) is the same
two-column shape the existing generator's batch tab already imports, so the
user can drag this file straight in.
"""

from __future__ import annotations

import datetime as _dt
import logging
import time
from pathlib import Path
from typing import Iterable

from openpyxl import Workbook, load_workbook

from crawler.fetcher import AmazonSession
from crawler.pipeline import S_OK, ProductResult, process_one

logger = logging.getLogger(__name__)


_OUT_XLSX_HEADERS = ("ZIP/文件夹路径", "商品标题")
_OUT_XLSX_NAME = "商品列表.xlsx"
_LOG_NAME = "crawl_log.txt"


def _looks_like_url(s: str | None) -> bool:
    if not s or not isinstance(s, str):
        return False
    s = s.strip()
    if "://" in s:
        return True
    # Accept bare ASINs like "B0GK22CBCR" too (10-char alphanumeric starting with B0)
    return len(s) == 10 and s[:2].upper() == "B0" and s.isalnum()


def read_urls_from_xlsx(input_xlsx: Path) -> list[str]:
    """Pull column A as a list of trimmed URL strings, auto-skipping a
    header row if present. Raises FileNotFoundError if the file is missing
    and ValueError if the workbook has no rows."""
    if not input_xlsx.is_file():
        raise FileNotFoundError(f"Input xlsx not found: {input_xlsx}")
    wb = load_workbook(filename=str(input_xlsx), read_only=True, data_only=True)
    try:
        ws = wb.active
        rows: list[str] = []
        for i, row in enumerate(ws.iter_rows(min_col=1, max_col=1, values_only=True)):
            v = row[0] if row else None
            if v is None:
                continue
            s = str(v).strip()
            if not s:
                continue
            if i == 0 and not _looks_like_url(s):
                # treat first non-empty cell as a header if it doesn't look like a URL
                continue
            rows.append(s)
    finally:
        wb.close()
    return rows


def _format_log_line(idx: int, total: int, r: ProductResult) -> str:
    ts = _dt.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    asin = r.asin or "????"
    title_excerpt = (r.title or "").replace("\n", " ").replace("\r", " ")[:80]
    return (
        f"[{ts}] [{idx}/{total}] {r.status:<12} {asin}  "
        f"imgs={r.image_count}  zip={r.zip_path or '-'}  "
        f"err={r.error}  title={title_excerpt}\n"
    )


def _format_console_line(idx: int, total: int, r: ProductResult) -> str:
    asin = r.asin or "----------"
    title_excerpt = (r.title or "").replace("\n", " ")[:50]
    return (
        f"[{idx:>3}/{total}] {r.status.upper():<11} {asin} | "
        f"{r.image_count:>2} imgs | {title_excerpt}"
    )


def _write_xlsx(out_path: Path, rows: Iterable[ProductResult]) -> int:
    """Write only ok rows to a 2-column xlsx; return the row count."""
    wb = Workbook()
    ws = wb.active
    ws.title = "商品列表"
    ws.cell(1, 1, _OUT_XLSX_HEADERS[0])
    ws.cell(1, 2, _OUT_XLSX_HEADERS[1])
    n = 0
    for r in rows:
        if r.status != S_OK or r.zip_path is None:
            continue
        ws.cell(n + 2, 1, str(r.zip_path))
        ws.cell(n + 2, 2, r.title)
        n += 1
    ws.column_dimensions["A"].width = 80
    ws.column_dimensions["B"].width = 60
    wb.save(str(out_path))
    return n


def run_batch(
    input_xlsx: Path,
    out_dir: Path,
    *,
    max_images: int = 8,
    limit: int | None = None,
    skip_existing: bool = False,
    delay_min: float = 2.0,
    delay_max: float = 5.0,
) -> dict[str, int]:
    """Drive the batch end-to-end. Returns a counts-by-status summary dict."""
    out_dir = Path(out_dir).resolve()
    out_dir.mkdir(parents=True, exist_ok=True)

    log_path = out_dir / _LOG_NAME
    xlsx_path = out_dir / _OUT_XLSX_NAME

    urls = read_urls_from_xlsx(Path(input_xlsx).resolve())
    if limit is not None and limit > 0:
        urls = urls[:limit]

    total = len(urls)
    if total == 0:
        print(f"No URLs found in {input_xlsx} (column A is empty).")
        return {"ok": 0}

    print(f"Crawling {total} URL(s) -> {out_dir}")
    print(
        f"  delay {delay_min:.1f}-{delay_max:.1f}s | max-images={max_images} | "
        f"skip-existing={skip_existing}"
    )

    # Session header at top of crawl_log so the user can tell runs apart.
    with log_path.open("a", encoding="utf-8") as fh:
        fh.write(
            f"\n=== Run @ {_dt.datetime.now():%Y-%m-%d %H:%M:%S} | total={total} | "
            f"out={out_dir} ===\n"
        )

    counts: dict[str, int] = {}
    results: list[ProductResult] = []
    started = time.monotonic()

    with AmazonSession(
        delay_min=delay_min,
        delay_max=delay_max,
        max_retries=3,
    ) as session:
        for i, url in enumerate(urls, start=1):
            r = process_one(
                url,
                session,
                out_dir,
                max_images=max_images,
                skip_existing=skip_existing,
            )
            results.append(r)
            counts[r.status] = counts.get(r.status, 0) + 1

            print(_format_console_line(i, total, r))
            with log_path.open("a", encoding="utf-8") as fh:
                fh.write(_format_log_line(i, total, r))

    n_in_xlsx = _write_xlsx(xlsx_path, results)
    elapsed = time.monotonic() - started

    summary = ", ".join(f"{k}={v}" for k, v in sorted(counts.items()))
    print(
        f"\nDone in {elapsed:.1f}s. {summary}\n"
        f"  -> {xlsx_path}  ({n_in_xlsx} ok rows)\n"
        f"  -> {log_path}"
    )
    return counts

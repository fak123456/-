"""Byte-level diagnostic: compare row 1's path (which succeeded) against
row 2's path (which failed) in the user's tasklist xlsx, to detect
invisible characters (BOM, zero-width space, trailing whitespace,
NBSP, etc.) that ``Path.is_file()`` rejects but the human eye misses.
"""
from __future__ import annotations

import sys
import unicodedata
from pathlib import Path

from openpyxl import load_workbook


def _describe_char(c: str) -> str:
    cp = ord(c)
    name = unicodedata.name(c, "?")
    cat = unicodedata.category(c)
    return f"U+{cp:04X} [{cat}] {name!r}"


def _annotate(s: str) -> str:
    """Return string with each char + its codepoint, flagging weird ones."""
    parts = []
    for c in s:
        cp = ord(c)
        weird = (
            cp < 0x20 or cp == 0x7F or
            cp in {0x00A0, 0x200B, 0x200C, 0x200D, 0xFEFF}
            or unicodedata.category(c) in {"Cf", "Mn", "Cs"}
        )
        if weird:
            parts.append(f"⚠️[{_describe_char(c)}]")
        elif c == " ":
            parts.append("·")
        else:
            parts.append(c)
    return "".join(parts)


def main() -> int:
    if len(sys.argv) < 2:
        print("usage: python _byte_diag_xlsx.py <path>")
        return 2
    p = Path(sys.argv[1])
    if not p.is_file():
        print(f"NOT FOUND: {p}")
        return 1

    wb = load_workbook(str(p), data_only=True, read_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    if len(rows) < 3:
        print("xlsx 只有", len(rows), "行 (含表头)，无可比对")
        return 1

    # Row 1 = HEADERS, Row 2 = first data row, Row 3 = second data row
    path1 = str(rows[1][0] or "")
    path2 = str(rows[2][0] or "")

    print(f"== 第 1 行（行号在 xlsx 中=2）==")
    print(f"  长度: {len(path1)}  | 字节数(UTF-8): {len(path1.encode('utf-8'))}")
    print(f"  raw : {path1!r}")
    print(f"  注解: {_annotate(path1)}")
    print()
    print(f"== 第 2 行（行号在 xlsx 中=3）==")
    print(f"  长度: {len(path2)}  | 字节数(UTF-8): {len(path2.encode('utf-8'))}")
    print(f"  raw : {path2!r}")
    print(f"  注解: {_annotate(path2)}")
    print()

    # Find common prefix length
    common = 0
    for a, b in zip(path1, path2):
        if a == b:
            common += 1
        else:
            break
    print(f"== 共同前缀长度: {common} ==")
    if common < min(len(path1), len(path2)):
        a_rest = path1[common:]
        b_rest = path2[common:]
        print(f"  第1行不同部分起始: {a_rest!r}  ({_annotate(a_rest)})")
        print(f"  第2行不同部分起始: {b_rest!r}  ({_annotate(b_rest)})")
    print()

    # Check Path.is_file() on both
    print("== 实际验证 ==")
    for label, s in [("第1行", path1), ("第2行", path2)]:
        rp = Path(s).expanduser().resolve()
        is_file = rp.is_file()
        is_dir = rp.is_dir()
        suffix = rp.suffix.lower()
        print(f"  {label}: is_file={is_file}  is_dir={is_dir}  suffix={suffix!r}  resolved={str(rp)[:80]!r}")
    print()

    # Check parent dir of row 1 — are there siblings?
    parent1 = Path(path1).expanduser().resolve().parent
    print(f"== 第1行所在目录: {parent1} ==")
    if parent1.is_dir():
        zips = sorted(parent1.glob("*.zip"))
        print(f"   该目录下 .zip 总数: {len(zips)}")
        # Look for the row-2 zip's bare name
        target = Path(path2).name
        match = [z for z in zips if z.name == target]
        if match:
            print(f"   ✅ 在该目录下能找到名字与第2行完全一致的 zip: {match[0]}")
        else:
            print(f"   ❌ 该目录下找不到文件名 == {target!r} 的 zip")
            # Look for similarly-named zips
            target_stem = Path(target).stem.strip()
            similar = [z for z in zips if target_stem in z.stem]
            if similar:
                print(f"   但找到 {len(similar)} 个名字相近的:")
                for z in similar[:5]:
                    print(f"     {z.name!r}  ({_annotate(z.name)})")
    else:
        print(f"   该目录不存在（在当前这台电脑上）")

    return 0


if __name__ == "__main__":
    raise SystemExit(main())

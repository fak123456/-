"""Read a user-supplied tasklist xlsx and dump every cell so we can see
exactly what they're seeing on screen.

We bypass ``load_tasklist`` first (raw cell dump), then also run
``load_tasklist`` to compare, because the user's complaint is usually
"the GUI shows X but my Excel shows Y" and the discrepancy lives in
the dataclass conversion or in trailing-blank-row handling.
"""
from __future__ import annotations

import sys
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT))

from gui.batch_tasklist import HEADERS, load_tasklist  # noqa: E402


def main() -> int:
    if len(sys.argv) < 2:
        print("usage: python _dump_user_xlsx.py <path>")
        return 2
    p = Path(sys.argv[1])
    if not p.is_file():
        print(f"NOT FOUND: {p}")
        return 1

    print(f"== {p}  ({p.stat().st_size} bytes) ==\n")

    print("---- 原始 cell dump (前 30 行) ----")
    wb = load_workbook(str(p), data_only=True, read_only=True)
    ws = wb.active
    print(f"sheet: {ws.title!r}  max_row={ws.max_row}  max_col={ws.max_column}")
    for ri, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if ri > 30:
            print("    … (剩余行省略)")
            break
        print(f"  R{ri:>3}: {row}")
    wb.close()

    print("\n---- load_tasklist 解析后的 TaskRow ----")
    rows = load_tasklist(p)
    print(f"  共 {len(rows)} 行 (TaskRow)")
    print(f"  HEADERS={HEADERS}")
    for i, r in enumerate(rows, 1):
        print(f"  第{i}行: status={r.status!r}  path={r.path[:80]!r}")
        print(f"          title={r.title[:60]!r}")
        if r.processed_at:
            print(f"          processed_at={r.processed_at}")
        if r.note:
            print(f"          note={r.note[:120]!r}")

    return 0


if __name__ == "__main__":
    raise SystemExit(main())

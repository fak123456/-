"""
Build a single-file Windows exe for the standalone Amazon crawler.

Use a clean venv that has only ``requests``, ``lxml``, ``openpyxl`` (and
``PyInstaller``) installed so the resulting exe stays small.

Usage (from repo root, with .venv-build active):
    python build_crawler_exe.py

Output:
    dist\\amazon_crawler.exe
    crawler_installer\\
        amazon_crawler.exe
        \u8f93\u5165\u6a21\u677f.xlsx
        \u4f7f\u7528\u8bf4\u660e.txt
        \u4e00\u952e\u751f\u6210\u8f93\u5165\u6a21\u677f.bat
        output\\        (empty placeholder)
    crawler_installer.zip       (the file you ship to others)
"""

from __future__ import annotations

import os
import shutil
import subprocess
import sys
import time
import zipfile
from pathlib import Path

ROOT = Path(__file__).resolve().parent
SPEC = ROOT / "crawler.spec"
EXE_NAME = "amazon_crawler.exe"
INSTALLER_DIR = ROOT / "crawler_installer"
ZIP_PATH = ROOT / "crawler_installer.zip"


def _stop_running_exe() -> None:
    if sys.platform != "win32":
        return
    subprocess.run(
        ["taskkill", "/F", "/IM", EXE_NAME],
        cwd=str(ROOT),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
    )
    time.sleep(1.0)


def _build_exe() -> Path:
    if not SPEC.is_file():
        raise SystemExit(f"Missing spec: {SPEC}")

    _stop_running_exe()

    dist = ROOT / "dist"
    build = ROOT / "build"
    if dist.is_dir():
        for p in dist.iterdir():
            if p.name == EXE_NAME or p.is_dir() and p.name == "amazon_crawler":
                if p.is_file():
                    p.unlink(missing_ok=True)
                else:
                    shutil.rmtree(p, ignore_errors=True)
    if (build / "amazon_crawler").is_dir():
        shutil.rmtree(build / "amazon_crawler", ignore_errors=True)

    cmd = [sys.executable, "-m", "PyInstaller", "--noconfirm", "--clean", str(SPEC)]
    print("Running:", " ".join(cmd))
    env = os.environ.copy()
    subprocess.check_call(cmd, cwd=str(ROOT), env=env)

    exe = ROOT / "dist" / EXE_NAME
    if not exe.is_file():
        raise SystemExit(f"Build finished but exe not found: {exe}")
    mb = exe.stat().st_size / (1024 * 1024)
    print(f"Built: {exe} ({mb:.1f} MB)")
    return exe


def _write_template(path: Path) -> None:
    """Use openpyxl directly to avoid invoking the freshly-built exe."""
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = "URLs"
    ws.cell(1, 1, "\u5546\u54c1\u94fe\u63a5")
    ws.cell(2, 1, "https://www.amazon.de/dp/B0CHX3QBCH")
    ws.column_dimensions["A"].width = 70
    wb.save(str(path))


def _write_readme(path: Path) -> None:
    text = (
        "\u4e9a\u9a6c\u900a\u5546\u54c1\u722c\u866b\uff08\u72ec\u7acb\u7248\uff09\n"
        "==========================\n"
        "\n"
        "\u3010\u4f5c\u7528\u3011\n"
        "\u8f93\u5165\u4e00\u4e2a Excel\u3010A \u5217\u586b\u4e9a\u9a6c\u900a\u5546\u54c1\u94fe\u63a5\u3011\uff0c"
        "\u811a\u672c\u4f1a\u4e3a\u6bcf\u4e2a\u5546\u54c1\u751f\u6210\u4e00\u4e2a {ASIN}.zip\uff08\u91cc\u9762\u662f\u6807\u9898 + \u5546\u54c1\u56fe\uff09\uff0c"
        "\u5e76\u8f93\u51fa\u4e00\u4e2a \u5546\u54c1\u5217\u8868.xlsx\uff0c\u53ef\u4ee5\u76f4\u63a5\u62d6\u5230\u3010\u7535\u5546\u751f\u56fe\u8f6f\u4ef6\u300d\u7684\u300c\u6279\u91cf\u300d\u9875\u9762\u4f7f\u7528\u3002\n"
        "\n"
        "\u3010\u4f7f\u7528\u6b65\u9aa4\u3011\n"
        "1\u3001\u6253\u5f00 \u8f93\u5165\u6a21\u677f.xlsx\uff0c\u628a\u4f60\u8981\u722c\u7684\u5546\u54c1\u94fe\u63a5\u4e00\u884c\u4e00\u4e2a\u7c98\u8d34\u5230 A \u5217\uff0c\u4fdd\u5b58\u3002\n"
        "   \uff08\u53ea\u8981 A \u5217\u91cc\u662f https://www.amazon.* \u5f00\u5934\u7684\u94fe\u63a5\u5c31\u884c\uff0c\u8868\u5934\u53ef\u4ee5\u4e0d\u8981\uff09\n"
        "\n"
        "2\u3001\u53cc\u51fb \u8fd0\u884c\u722c\u866b.bat\u3002\n"
        "   \u4f1a\u5f39\u4e2a\u9ed1\u7a97\u53e3\u5f00\u59cb\u8dd1\uff0c\u4e00\u4e2a\u5546\u54c1\u5927\u7ea6 5\uff5e10 \u79d2\u3002\n"
        "   \u4e2d\u9014\u4e0d\u8981\u5173\u9ed1\u7a97\u53e3\uff01\u8dd1\u5b8c\u4f1a\u63d0\u793a\u300cPress any key to continue\u300d\u3002\n"
        "\n"
        "3\u3001\u53bb output \u6587\u4ef6\u5939\u62ff\u7ed3\u679c\uff1a\n"
        "      output\\B0CHX3QBCH.zip       \u2190 \u6bcf\u4e2a\u5546\u54c1\u4e00\u4e2a zip\n"
        "      output\\B0XXXXXXXX.zip\n"
        "      output\\\u5546\u54c1\u5217\u8868.xlsx       \u2190 \u62d6\u8fdb\u751f\u56fe\u8f6f\u4ef6\u7528\n"
        "      output\\crawl_log.txt        \u2190 \u8be6\u7ec6\u65e5\u5fd7\uff08\u51fa\u95ee\u9898\u624d\u770b\uff09\n"
        "\n"
        "\u3010\u8d77\u4e0d\u6765\u5982\u4f55\u624b\u52a8\u8dd1\u3011\n"
        "  \u6253\u5f00\u672c\u6587\u4ef6\u5939\uff0c\u5728\u5730\u5740\u680f\u8f93 cmd \u56de\u8f66\uff0c\u7136\u540e\u8d34\uff1a\n"
        "      amazon_crawler.exe --input \u8f93\u5165\u6a21\u677f.xlsx --out output\n"
        "  \u53ef\u9009\u53c2\u6570\uff1a\n"
        "      --max-images 8        \u6bcf\u4ef6\u5546\u54c1\u6700\u591a\u62b9\u51e0\u5f20\u56fe\uff08\u9ed8\u8ba4 8\uff09\n"
        "      --skip-existing       \u8df3\u8fc7\u5df2\u7ecf\u6709 zip \u7684 ASIN\uff0c\u65ad\u70b9\u91cd\u8dd1\u7528\n"
        "      --delay-min 2 --delay-max 5    \u6bcf\u9875\u4e4b\u95f4\u968f\u673a\u5ef6\u8fdf 2\uff5e5 \u79d2\uff08\u9ed8\u8ba4\u503c\uff09\n"
        "      --limit 5             \u53ea\u8dd1\u524d 5 \u4e2a\uff0c\u8c03\u8bd5\u7528\n"
        "\n"
        "\u3010\u5e38\u89c1\u95ee\u9898\u3011\n"
        "Q\uff1a\u67d0\u4e2a\u5546\u54c1 FETCH_ERROR / 404 / CAPTCHA \u600e\u4e48\u529e\uff1f\n"
        "A\uff1a\u4ed6\u4f1a\u8df3\u8fc7\u8fd9\u4e2a\u5546\u54c1\u7ee7\u7eed\u8dd1\uff0c\u4e0d\u4f1a\u7838\u6574\u4e2a\u4efb\u52a1\u3002\u8be6\u60c5\u770b crawl_log.txt\u3002\n"
        "    \u88abrobot-check\uff08CAPTCHA\uff09\u62e6\u4f4f\uff1a\u7b49\u5341\u51e0\u5206\u949f\u540d\u8005\u6362\u7f51\u7edc\u91cd\u8dd1\uff0c\u4e0d\u8981\u540c\u4e00\u5206\u949f\u91cd\u8d77 N \u6b21\u3002\n"
        "    \u4ea7\u54c1 404\uff1a\u5546\u54c1\u672c\u8eab\u5728\u8be5\u7ad9\u70b9\u4e0a\u4e0b\u67b6\u4e86\uff0c\u68c0\u67e5\u94fe\u63a5\u3002\n"
        "\n"
        "Q\uff1a\u8dd1\u8d77\u6765\u63d0\u793a\u300c\u65e0\u6cd5\u8bbf\u95ee\u8be5\u7f51\u7ad9\u300d\u600e\u4e48\u529e\uff1f\n"
        "A\uff1a\u9700\u8981\u80fd\u8bbf\u95ee\u4e9a\u9a6c\u900a\u7684\u7f51\u7edc\uff08VPN / \u4e13\u7ebf\uff09\u3002\n"
        "\n"
        "Q\uff1a\u4e00\u6b21\u80fd\u8dd1\u591a\u5c11\u5546\u54c1\uff1f\n"
        "A\uff1a\u5b9e\u9645\u4e0a\u4e9a\u9a6c\u900a\u7eaf\u9762\u5c0f\u65f6\u8bbf\u95ee\u4e0a\u9650\u5728\u51e0\u767e\u6761\u91cf\u7ea7\uff0c\u8d85\u4e86\u4f1a\u88ab\u9650\u6d41\u3002\u5efa\u8bae\u6bcf\u6279\u22641000 \u4ef6\u3002\n"
        "\n"
        "\u3010\u6cd5\u5f8b\u63d0\u793a\u3011\n"
        "\u672c\u811a\u672c\u4ec5\u4f9b\u4e2a\u4eba\u5b66\u4e60\u4e0e\u8fd0\u8425\u4f7f\u7528\u3002\u4e9a\u9a6c\u900a\u670d\u52a1\u6761\u6b3e\u4e0d\u5141\u8bb8\u81ea\u52a8\u5316\u8bbf\u95ee\u3002"
        "\u8bf7\u4e0d\u8981\u9ad8\u9891\u8bf7\u6c42\uff0c\u4e5f\u4e0d\u8981\u7528\u4e8e\u5546\u4e1a\u91cd\u53d1\u522b\u4eba\u7684\u56fe\u7247 / \u6587\u6848\u3002\n"
    )
    path.write_text(text, encoding="utf-8")


def _write_run_bat(path: Path) -> None:
    text = (
        "@echo off\r\n"
        "chcp 65001 >nul\r\n"
        "cd /d \"%~dp0\"\r\n"
        "if not exist output mkdir output\r\n"
        "echo \u5f00\u59cb\u722c\u53d6\uff0c\u8fc7\u7a0b\u4e2d\u4e0d\u8981\u5173\u95ed\u672c\u7a97\u53e3\u3002\r\n"
        "echo.\r\n"
        "amazon_crawler.exe --input \"\u8f93\u5165\u6a21\u677f.xlsx\" --out output\r\n"
        "echo.\r\n"
        "echo \u5b8c\u6210\u3002\u8bf7\u53bb output \u6587\u4ef6\u5939\u67e5\u770b\u7ed3\u679c\u3002\r\n"
        "pause\r\n"
    )
    path.write_text(text, encoding="utf-8")


def _write_make_template_bat(path: Path) -> None:
    text = (
        "@echo off\r\n"
        "chcp 65001 >nul\r\n"
        "cd /d \"%~dp0\"\r\n"
        "amazon_crawler.exe --input \"\u8f93\u5165\u6a21\u677f.xlsx\" --make-template\r\n"
        "echo.\r\n"
        "echo \u5df2\u91cd\u65b0\u751f\u6210\u7a7a\u8f93\u5165\u6a21\u677f.xlsx\u3002\r\n"
        "pause\r\n"
    )
    path.write_text(text, encoding="utf-8")


def _build_installer(exe_path: Path) -> Path:
    if INSTALLER_DIR.is_dir():
        shutil.rmtree(INSTALLER_DIR, ignore_errors=True)
    INSTALLER_DIR.mkdir(parents=True)

    shutil.copy2(exe_path, INSTALLER_DIR / EXE_NAME)
    _write_template(INSTALLER_DIR / "\u8f93\u5165\u6a21\u677f.xlsx")
    _write_readme(INSTALLER_DIR / "\u4f7f\u7528\u8bf4\u660e.txt")
    _write_run_bat(INSTALLER_DIR / "\u8fd0\u884c\u722c\u866b.bat")
    _write_make_template_bat(INSTALLER_DIR / "\u91cd\u751f\u6210\u8f93\u5165\u6a21\u677f.bat")
    (INSTALLER_DIR / "output").mkdir(exist_ok=True)
    (INSTALLER_DIR / "output" / ".keep").write_text("", encoding="utf-8")

    print(f"Installer dir ready: {INSTALLER_DIR}")
    return INSTALLER_DIR


def _zip_installer(installer_dir: Path) -> Path:
    if ZIP_PATH.is_file():
        ZIP_PATH.unlink()
    with zipfile.ZipFile(ZIP_PATH, "w", zipfile.ZIP_DEFLATED) as zf:
        for f in installer_dir.rglob("*"):
            if f.is_file():
                zf.write(f, arcname=f.relative_to(installer_dir.parent))
    mb = ZIP_PATH.stat().st_size / (1024 * 1024)
    print(f"Wrote: {ZIP_PATH} ({mb:.1f} MB)")
    return ZIP_PATH


def main() -> int:
    exe = _build_exe()
    installer = _build_installer(exe)
    _zip_installer(installer)
    print("Done.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())

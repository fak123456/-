"""Persistent batch task list (resume-on-crash for the 批量 tab).

Each instance has its own xlsx file (default: ``<exe_dir>/批量任务_实例<N>.xlsx``).
The user fills columns A/B (path + title); columns C/D/E are written by the
program after each row finishes (status, processed-at, note). The program
saves the xlsx after every row so the user can interrupt at any time and
re-running picks up exactly where they left off — already-done rows are
skipped without consuming any API quota.

Also tolerates uploads of legacy 2-column files: missing C/D/E columns are
treated as ``待处理`` for every row.
"""

from __future__ import annotations

import os
import re
import shutil
import tempfile
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook

# ---- Status constants (Chinese, shown directly in the xlsx + UI) ----------

STATUS_PENDING = "待处理"
STATUS_DONE = "已完成"
STATUS_PARTIAL = "部分完成"
STATUS_FAILED = "失败"
STATUS_CANCELLED = "已终止"  # user hit ⏹ mid-flight; meta has ``cancelled=True``
STATUS_RUNNING = "运行中"  # transient, written briefly while a row is in flight

ALL_STATUSES = {
    STATUS_PENDING,
    STATUS_DONE,
    STATUS_PARTIAL,
    STATUS_FAILED,
    STATUS_CANCELLED,
    STATUS_RUNNING,
}

HEADERS = ["ZIP/文件夹路径", "商品标题", "状态", "处理时间", "备注"]

# ---- TaskRow --------------------------------------------------------------


@dataclass
class TaskRow:
    """One product row in the task list."""

    path: str = ""
    title: str = ""
    status: str = STATUS_PENDING
    processed_at: str = ""
    note: str = ""

    def is_actionable(
        self,
        *,
        retry_failed: bool,
        retry_partial: bool,
        retry_cancelled: bool = True,
    ) -> bool:
        """Return ``True`` if this row should be processed in the current run.

        ``retry_cancelled`` defaults to True because the *expected* behaviour
        when the user hit ⏹ is "续跑没生成的图" — leaving it off means the row
        gets stuck. The runner uses ``compute_missing_slots`` to make sure
        续跑 is cheap (already-OK slots are not re-generated).
        """
        if not self.path.strip():
            return False
        if self.status == STATUS_DONE:
            return False
        if self.status == STATUS_FAILED and not retry_failed:
            return False
        if self.status == STATUS_PARTIAL and not retry_partial:
            return False
        if self.status == STATUS_CANCELLED and not retry_cancelled:
            return False
        # STATUS_PENDING, STATUS_RUNNING (orphaned), and "" all run.
        return True

    def to_row(self) -> list[str]:
        return [self.path, self.title, self.status, self.processed_at, self.note]


# ---- Path resolution ------------------------------------------------------


def _instance_label() -> str:
    """Return the per-instance suffix used in the default file name."""
    raw = os.environ.get("AMAZON_IMG_GUI_INSTANCE", "").strip()
    cleaned = re.sub(r"[^A-Za-z0-9_-]", "", raw)
    return cleaned or ""


def default_tasklist_path() -> Path:
    """Default xlsx path: beside the exe (or repo root in dev), with instance suffix."""
    from gui.paths import app_workdir

    label = _instance_label()
    name = f"批量任务_实例{label}.xlsx" if label else "批量任务.xlsx"
    return (app_workdir() / name).resolve()


# ---- Read / write ---------------------------------------------------------


def _now_str() -> str:
    return datetime.now().strftime("%Y-%m-%d %H:%M:%S")


def _normalize_status(s: str) -> str:
    s = (s or "").strip()
    if not s:
        return STATUS_PENDING
    if s in ALL_STATUSES:
        return s
    return STATUS_PENDING


def _looks_like_header(first_cell: str) -> bool:
    s = (first_cell or "").strip()
    if not s:
        return True
    s_lower = s.lower()
    needles = ("zip", "文件夹", "路径", "path", "url", "链接")
    return any(n in s_lower or n in s for n in needles)


def init_template_if_missing(path: Path) -> bool:
    """Create a 5-column template at ``path`` if it doesn't exist yet.

    Returns ``True`` if the file was created, ``False`` if it already existed.
    """
    if path.is_file():
        return False
    path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "任务"
    for col_idx, h in enumerate(HEADERS, start=1):
        ws.cell(1, col_idx, h)
    ws.cell(2, 1, "")
    ws.cell(2, 2, "")
    widths = [55, 50, 12, 22, 40]
    for col_idx, w in enumerate(widths, start=1):
        ws.column_dimensions[ws.cell(1, col_idx).column_letter].width = w
    wb.save(str(path))
    return True


def load_tasklist(path: Path) -> list[TaskRow]:
    """Read all rows from the xlsx. Auto-creates the file if missing.

    Trailing blank rows are dropped. Rows whose path column is empty are
    skipped silently (they're just spacer rows in the user's worksheet).
    Legacy 2-column files load fine — C/D/E default to ``待处理`` / ``""``.
    """
    if not path.is_file():
        init_template_if_missing(path)
        return []

    wb = load_workbook(str(path), data_only=True, read_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)

    rows: list[TaskRow] = []
    skipped_header = False
    for raw in rows_iter:
        if raw is None:
            continue
        cells = list(raw) + [None] * max(0, 5 - len(raw))
        c0 = "" if cells[0] is None else str(cells[0]).strip()
        c1 = "" if cells[1] is None else str(cells[1]).strip()
        c2 = "" if cells[2] is None else str(cells[2]).strip()
        c3 = "" if cells[3] is None else str(cells[3]).strip()
        c4 = "" if cells[4] is None else str(cells[4]).strip()

        if not skipped_header and _looks_like_header(c0):
            skipped_header = True
            continue

        if not c0 and not c1 and not c2:
            continue

        rows.append(
            TaskRow(
                path=c0,
                title=c1,
                status=_normalize_status(c2),
                processed_at=c3,
                note=c4,
            )
        )

    wb.close()
    return rows


def save_tasklist(path: Path, rows: list[TaskRow]) -> None:
    """Write all rows back to the xlsx atomically.

    Raises ``PermissionError`` if the target file is locked (Excel still has
    it open). Caller should surface a friendly "请先关闭 Excel 再点开始"
    message in that case.

    Writes to ``<name>.tmp`` then os.replace's it onto the target so a crash
    mid-write can never corrupt the previously-saved progress.
    """
    path.parent.mkdir(parents=True, exist_ok=True)
    fd, tmp_str = tempfile.mkstemp(prefix=path.stem + ".", suffix=".tmp", dir=str(path.parent))
    os.close(fd)
    tmp = Path(tmp_str)
    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "任务"
        for col_idx, h in enumerate(HEADERS, start=1):
            ws.cell(1, col_idx, h)
        for row_idx, r in enumerate(rows, start=2):
            for col_idx, val in enumerate(r.to_row(), start=1):
                ws.cell(row_idx, col_idx, val)
        widths = [55, 50, 12, 22, 40]
        for col_idx, w in enumerate(widths, start=1):
            ws.column_dimensions[ws.cell(1, col_idx).column_letter].width = w
        wb.save(str(tmp))
        # os.replace is atomic on Windows; raises PermissionError if dest is locked.
        os.replace(str(tmp), str(path))
    finally:
        if tmp.exists():
            try:
                tmp.unlink()
            except OSError:
                pass


def mark_row(
    rows: list[TaskRow],
    row_index: int,
    *,
    status: str,
    note: str | None = None,
) -> None:
    """Update one row's status / processed_at / note in-place."""
    if not 0 <= row_index < len(rows):
        return
    r = rows[row_index]
    r.status = status
    r.processed_at = _now_str()
    if note is not None:
        r.note = note


def reset_all_status(rows: list[TaskRow]) -> None:
    """Wipe C/D/E for every row (useful for "全部重置" button)."""
    for r in rows:
        r.status = STATUS_PENDING
        r.processed_at = ""
        r.note = ""


# ---- Display helpers ------------------------------------------------------


def to_df_rows(rows: list[TaskRow]) -> list[list[str]]:
    """Convert to the 2D list shape the gr.Dataframe component wants."""
    return [r.to_row() for r in rows]


def summarize(rows: list[TaskRow]) -> dict[str, int]:
    """Count rows per status. Useful for the live summary banner."""
    counts: dict[str, int] = {
        s: 0
        for s in (
            STATUS_PENDING,
            STATUS_DONE,
            STATUS_PARTIAL,
            STATUS_FAILED,
            STATUS_CANCELLED,
            STATUS_RUNNING,
        )
    }
    counts["total"] = len(rows)
    for r in rows:
        s = r.status if r.status in counts else STATUS_PENDING
        counts[s] = counts.get(s, 0) + 1
    return counts


def format_summary(rows: list[TaskRow]) -> str:
    """One-liner status summary suitable for a Markdown component."""
    if not rows:
        return "（任务清单为空）"
    c = summarize(rows)
    return (
        f"**总计 {c['total']}** ｜"
        f" 待处理 **{c[STATUS_PENDING]}** ｜"
        f" 已完成 **{c[STATUS_DONE]}** ｜"
        f" 部分完成 **{c[STATUS_PARTIAL]}** ｜"
        f" 已终止 **{c[STATUS_CANCELLED]}** ｜"
        f" 失败 **{c[STATUS_FAILED]}**"
    )


# ---- Open in system default editor ---------------------------------------


def open_in_default_app(path: Path) -> tuple[bool, str]:
    """Open the xlsx in Excel / system default app. Windows only meaningfully.

    Returns ``(ok, message)``.
    """
    if not path.is_file():
        return False, f"文件不存在: {path}"
    try:
        if hasattr(os, "startfile"):
            os.startfile(str(path))  # type: ignore[attr-defined]
            return True, f"已用系统默认程序打开: {path.name}"
        # Linux/macOS fallback (rarely used by the operations team but cheap to keep)
        import subprocess

        opener = "open" if shutil.which("open") else "xdg-open"
        subprocess.Popen([opener, str(path)])
        return True, f"已用 {opener} 打开: {path.name}"
    except Exception as e:
        return False, f"打开失败: {e}"

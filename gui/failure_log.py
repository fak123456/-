"""Persistent per-slot failure record for the 批量 tab.

Tracks every individual image slot that returned ``status=error`` during a
batch run, so the user can later click 「🔁 重跑全部失败图片」 and have just
those slots regenerated. Companion to ``gui/batch_tasklist.py``.

The xlsx lives next to the task-list xlsx, with the file name pattern
``批量失败记录_实例<N>.xlsx`` (or ``批量失败记录.xlsx`` when no instance is set).

Schema (7 columns, written / read by this module):

| A 商品文件夹 | B 槽位 | C 错误信息 | D 失败时间 | E 重跑状态 | F 重跑时间 | G 商品路径 |
|---|---|---|---|---|---|---|

槽位 follows the same TYPE_NN format expected by ``parse_regen_targets``,
e.g. ``main_01`` / ``scene_02``.
"""

from __future__ import annotations

import json
import os
import re
import tempfile
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Iterable

from openpyxl import Workbook, load_workbook

# ---- Status constants -----------------------------------------------------

RETRY_PENDING = "待重跑"
RETRY_OK = "已重跑成功"
RETRY_FAIL = "重跑仍失败"
RETRY_SKIPPED = "跳过"  # set when user manually marks a row as ignored

ALL_RETRY_STATUSES = {RETRY_PENDING, RETRY_OK, RETRY_FAIL, RETRY_SKIPPED}

HEADERS = [
    "商品文件夹",
    "槽位",
    "错误信息",
    "失败时间",
    "重跑状态",
    "重跑时间",
    "商品路径",
]

# Width of each column in the saved xlsx (visual-only; failure-tolerant).
_COL_WIDTHS = [22, 12, 50, 22, 12, 22, 60]

# ---- Slot key helpers -----------------------------------------------------

_SLOT_RE = re.compile(r"^([a-z]+)_(\d+)$")


def slot_label(type_name: str, idx: int) -> str:
    """Return the canonical regen-target slot label, e.g. ``main_01``."""
    return f"{type_name}_{int(idx):02d}"


def parse_slot(slot: str) -> tuple[str, int] | None:
    """Inverse of ``slot_label``. Returns ``None`` for malformed slots."""
    m = _SLOT_RE.fullmatch((slot or "").strip())
    if not m:
        return None
    return m.group(1), int(m.group(2))


# ---- FailureRow -----------------------------------------------------------


@dataclass
class FailureRow:
    """One failed image slot."""

    product_dir_name: str = ""
    slot: str = ""
    error: str = ""
    failed_at: str = ""
    retry_status: str = RETRY_PENDING
    retried_at: str = ""
    product_path: str = ""

    def key(self) -> tuple[str, str]:
        """Identity key for dedup ((product_path, slot) is unique)."""
        return (self.product_path, self.slot)

    def to_row(self) -> list[str]:
        return [
            self.product_dir_name,
            self.slot,
            self.error,
            self.failed_at,
            self.retry_status,
            self.retried_at,
            self.product_path,
        ]

    def is_retryable(self, *, include_already_retried: bool) -> bool:
        if not self.product_path or not self.slot:
            return False
        if self.retry_status == RETRY_SKIPPED:
            return False
        if self.retry_status == RETRY_OK and not include_already_retried:
            return False
        return True


# ---- Path resolution ------------------------------------------------------


def _instance_label() -> str:
    raw = os.environ.get("AMAZON_IMG_GUI_INSTANCE", "").strip()
    return re.sub(r"[^A-Za-z0-9_-]", "", raw)


def default_failure_log_path() -> Path:
    """Default xlsx path: next to the exe (or repo root in dev), instance suffix."""
    from gui.paths import app_workdir

    label = _instance_label()
    name = f"批量失败记录_实例{label}.xlsx" if label else "批量失败记录.xlsx"
    return (app_workdir() / name).resolve()


def derive_failure_log_path(tasklist_path: Path) -> Path:
    """Make a sibling failure-log xlsx next to the user's task-list xlsx.

    Prefers the ``批量任务 → 批量失败记录`` rename so an instance-customised
    task list keeps its instance label without us having to re-derive it.
    """
    tasklist_path = Path(tasklist_path)
    name = tasklist_path.name
    stem, ext = tasklist_path.stem, tasklist_path.suffix or ".xlsx"
    if name.startswith("批量任务"):
        new_name = name.replace("批量任务", "批量失败记录", 1)
    else:
        new_name = f"{stem}_失败记录{ext}"
    return tasklist_path.parent / new_name


# ---- Now string -----------------------------------------------------------


def _now_str() -> str:
    return datetime.now().strftime("%Y-%m-%d %H:%M:%S")


# ---- Read / write ---------------------------------------------------------


def _looks_like_header(first_cell: str) -> bool:
    s = (first_cell or "").strip()
    if not s:
        return True
    return any(n in s for n in ("商品", "文件夹", "ASIN", "Folder", "folder"))


def init_template_if_missing(path: Path) -> bool:
    """Create an empty xlsx with the standard 7-column header. Idempotent."""
    if path.is_file():
        return False
    path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "失败记录"
    for col_idx, h in enumerate(HEADERS, start=1):
        ws.cell(1, col_idx, h)
    for col_idx, w in enumerate(_COL_WIDTHS, start=1):
        ws.column_dimensions[ws.cell(1, col_idx).column_letter].width = w
    wb.save(str(path))
    return True


def load_failure_log(path: Path) -> list[FailureRow]:
    """Read all failure rows. Auto-creates the file if missing.

    Tolerates malformed rows (skips silently). Empty/blank rows are dropped.
    """
    if not path.is_file():
        init_template_if_missing(path)
        return []

    wb = load_workbook(str(path), data_only=True, read_only=True)
    ws = wb.active

    rows: list[FailureRow] = []
    skipped_header = False
    for raw in ws.iter_rows(values_only=True):
        if raw is None:
            continue
        cells = list(raw) + [None] * max(0, 7 - len(raw))
        c = ["" if x is None else str(x).strip() for x in cells[:7]]
        if not skipped_header and _looks_like_header(c[0]):
            skipped_header = True
            continue
        # Need at least slot + path to be useful
        if not c[1] and not c[6]:
            continue
        rs = c[4] if c[4] in ALL_RETRY_STATUSES else RETRY_PENDING
        rows.append(
            FailureRow(
                product_dir_name=c[0],
                slot=c[1],
                error=c[2],
                failed_at=c[3],
                retry_status=rs,
                retried_at=c[5],
                product_path=c[6],
            )
        )
    wb.close()
    return rows


def save_failure_log(path: Path, rows: list[FailureRow]) -> None:
    """Atomic write. Raises ``PermissionError`` if Excel still has it open."""
    path.parent.mkdir(parents=True, exist_ok=True)
    fd, tmp_str = tempfile.mkstemp(prefix=path.stem + ".", suffix=".tmp", dir=str(path.parent))
    os.close(fd)
    tmp = Path(tmp_str)
    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "失败记录"
        for col_idx, h in enumerate(HEADERS, start=1):
            ws.cell(1, col_idx, h)
        for row_idx, r in enumerate(rows, start=2):
            for col_idx, val in enumerate(r.to_row(), start=1):
                ws.cell(row_idx, col_idx, val)
        for col_idx, w in enumerate(_COL_WIDTHS, start=1):
            ws.column_dimensions[ws.cell(1, col_idx).column_letter].width = w
        wb.save(str(tmp))
        os.replace(str(tmp), str(path))
    finally:
        if tmp.exists():
            try:
                tmp.unlink()
            except OSError:
                pass


# ---- Mutations ------------------------------------------------------------


def merge_failures_from_meta(
    rows: list[FailureRow],
    *,
    product_dir: Path,
    meta: dict | None,
) -> int:
    """Reconcile the failure log with this product's current ``meta.json``.

    Semantics — the failure log is "still-failing slots that need attention":

    * **status=ok** in meta → if a matching row exists, **drop it** (slot is
      now fixed, no reason to keep it in the open-issue list).
    * **status=error** in meta → if a matching row exists, refresh its
      ``error`` / ``failed_at`` (and reset retry state to 待重跑 so a new
      retry pass picks it up). Otherwise append a fresh row.

    Returns the count of rows touched (added / refreshed / dropped).
    """
    if not isinstance(meta, dict):
        return 0
    images = meta.get("images") or []
    if not isinstance(images, list):
        return 0

    pdir_path = str(product_dir.resolve())
    pdir_name = product_dir.name

    # First pass: bucket meta into ok / error slots.
    ok_slots: set[str] = set()
    err_slots: dict[str, str] = {}
    for im in images:
        if not isinstance(im, dict):
            continue
        type_name = str(im.get("type") or "").strip()
        try:
            idx = int(im.get("index") or 0)
        except (TypeError, ValueError):
            continue
        if not type_name or idx <= 0:
            continue
        slot = slot_label(type_name, idx)
        status = str(im.get("status") or "").strip()
        if status == "ok":
            ok_slots.add(slot)
        elif status == "error":
            err_slots[slot] = str(im.get("error") or "").strip() or "(no error message)"

    # Drop rows for this product whose slot is now ok.
    n_removed = 0
    if ok_slots:
        before = len(rows)
        rows[:] = [
            r for r in rows
            if not (r.product_path == pdir_path and r.slot in ok_slots)
        ]
        n_removed = before - len(rows)

    # Refresh existing error rows / append new ones.
    by_key: dict[tuple[str, str], FailureRow] = {r.key(): r for r in rows}
    now = _now_str()
    touched = n_removed
    for slot, err in err_slots.items():
        key = (pdir_path, slot)
        existing = by_key.get(key)
        if existing is not None:
            existing.product_dir_name = pdir_name
            existing.error = err
            existing.failed_at = existing.failed_at or now
            # Refreshed via a fresh failure → re-arm for the next retry pass.
            existing.retry_status = RETRY_PENDING
            existing.retried_at = ""
        else:
            new_row = FailureRow(
                product_dir_name=pdir_name,
                slot=slot,
                error=err,
                failed_at=now,
                retry_status=RETRY_PENDING,
                retried_at="",
                product_path=pdir_path,
            )
            rows.append(new_row)
            by_key[key] = new_row
        touched += 1

    return touched


def mark_just_retried_failures(
    rows: list[FailureRow],
    *,
    product_dir: Path,
    retried_slots: set[str],
) -> int:
    """Mark rows for ``product_dir`` whose slot is in ``retried_slots`` as
    ``重跑仍失败``.

    Call this *after* ``merge_failures_from_meta`` during a retry pass — any
    slot that was successfully fixed has already been removed from ``rows``,
    so what's left for the same product is by definition "we tried, it's
    still bad", which is what the user wants flagged in the xlsx.

    Returns the number of rows updated.
    """
    pdir_path = str(product_dir.resolve())
    now = _now_str()
    n = 0
    for r in rows:
        if r.product_path != pdir_path:
            continue
        if r.slot not in retried_slots:
            continue
        r.retry_status = RETRY_FAIL
        r.retried_at = now
        n += 1
    return n


def clear_resolved(rows: list[FailureRow]) -> int:
    """Drop rows whose retry_status is 已重跑成功 (or 跳过). Returns dropped count.

    Mostly redundant after the auto-drop semantics in
    ``merge_failures_from_meta`` — kept for backward compat and for cleaning
    up legacy xlsx files that were written under the old behaviour.
    """
    keep: list[FailureRow] = []
    dropped = 0
    for r in rows:
        if r.retry_status in (RETRY_OK, RETRY_SKIPPED):
            dropped += 1
            continue
        keep.append(r)
    rows[:] = keep
    return dropped


def clear_all(rows: list[FailureRow]) -> int:
    """Wipe every row (nuclear option for the ♻️ 清空全部失败记录 button).

    Returns the count of rows removed so the UI can show "已清掉 N 条".
    """
    n = len(rows)
    rows.clear()
    return n


def reset_all_to_pending(rows: list[FailureRow]) -> None:
    """Force every row back to 待重跑 (used by the manual ♻️ button)."""
    for r in rows:
        if r.retry_status == RETRY_OK:
            continue
        r.retry_status = RETRY_PENDING
        r.retried_at = ""


# ---- Display helpers ------------------------------------------------------


def to_df_rows(rows: list[FailureRow]) -> list[list[str]]:
    return [r.to_row() for r in rows]


def summarize(rows: list[FailureRow]) -> dict[str, int]:
    counts = {s: 0 for s in (RETRY_PENDING, RETRY_OK, RETRY_FAIL, RETRY_SKIPPED)}
    counts["total"] = len(rows)
    for r in rows:
        s = r.retry_status if r.retry_status in counts else RETRY_PENDING
        counts[s] = counts.get(s, 0) + 1
    return counts


def format_summary(rows: list[FailureRow]) -> str:
    if not rows:
        return "（暂无失败记录）"
    c = summarize(rows)
    return (
        f"**失败记录 {c['total']}** ｜"
        f" 待重跑 **{c[RETRY_PENDING]}** ｜"
        f" 已重跑成功 **{c[RETRY_OK]}** ｜"
        f" 重跑仍失败 **{c[RETRY_FAIL]}**"
    )


def group_by_product(
    rows: list[FailureRow],
    *,
    include_already_retried: bool = False,
) -> list[tuple[Path, str, list[tuple[int, FailureRow]]]]:
    """Group retryable failure rows by product path.

    Returns a list of ``(product_path, product_dir_name, [(row_index, FailureRow), ...])``
    tuples. The row indices are positions inside the original ``rows`` list,
    so callers can update them in place after a retry.
    """
    out_map: dict[str, list[tuple[int, FailureRow]]] = {}
    name_map: dict[str, str] = {}
    for i, r in enumerate(rows):
        if not r.is_retryable(include_already_retried=include_already_retried):
            continue
        out_map.setdefault(r.product_path, []).append((i, r))
        name_map[r.product_path] = r.product_dir_name
    out: list[tuple[Path, str, list[tuple[int, FailureRow]]]] = []
    for path_str, items in out_map.items():
        out.append((Path(path_str), name_map.get(path_str, ""), items))
    return out


def open_in_default_app(path: Path) -> tuple[bool, str]:
    """Open the xlsx in Excel / system default app."""
    if not path.is_file():
        return False, f"文件不存在: {path}"
    try:
        if hasattr(os, "startfile"):
            os.startfile(str(path))  # type: ignore[attr-defined]
            return True, f"已用系统默认程序打开: {path.name}"
        import shutil
        import subprocess

        opener = "open" if shutil.which("open") else "xdg-open"
        subprocess.Popen([opener, str(path)])
        return True, f"已用 {opener} 打开: {path.name}"
    except Exception as e:
        return False, f"打开失败: {e}"


def read_meta(product_dir: Path) -> dict | None:
    """Convenience reader for ``output/meta.json``. Returns ``None`` if absent."""
    p = product_dir / "output" / "meta.json"
    if not p.is_file():
        return None
    try:
        return json.loads(p.read_text(encoding="utf-8"))
    except Exception:
        return None
